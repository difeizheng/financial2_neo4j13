"""Table boundary detection within Excel sheets.

Each sheet may contain one or more logical tables. This module uses
spatial analysis adapted from excel_table_extractor.py:

1. Identify vertical merged cells (>=3 rows) as L-shaped table anchors
2. For cells NOT in any anchor's territory, BFS-expand as rectangular tables
3. For each anchor, expand an L-table via row-scanning (not cell-level BFS)
   — this is the key insight: row-scanning tolerates sparse columns within a table
4. Merge adjacent tables on the same rows with small column gaps
5. Classify columns by header labels and data sampling
6. Filter out trivial tables (<=1 row or <=1 cell)
"""
from __future__ import annotations
import re
from datetime import datetime
from typing import Optional

from openpyxl.utils import get_column_letter, column_index_from_string


# --- Column role keywords ---

_UNIT_KEYWORDS = {"单位", "万元", "元", "亿元", "%", "MW", "kW", "h", "年", "月", "个"}
_NAME_KEYWORDS = {"项目", "名称", "参数", "指标", "科目"}
_SEQ_KEYWORDS = {"序号", "编号", "序"}
_CATEGORY_KEYWORDS = {"类别", "分类", "大类"}
_TOTAL_KEYWORDS = {"合计", "小计", "汇总", "总计", "总额"}
_NOTES_KEYWORDS = {"备注", "说明", "注", "取值说明", "参数释义"}
_HEADER_KEYWORDS = _SEQ_KEYWORDS | _NAME_KEYWORDS | _CATEGORY_KEYWORDS | _UNIT_KEYWORDS | _NOTES_KEYWORDS


def _excel_serial_to_label(serial: float) -> str:
    """Convert Excel date serial number to standardized label."""
    try:
        dt = datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(serial))
        return dt.strftime("%Y-%m")
    except Exception:
        return str(int(serial))


def _is_time_value(v) -> tuple[bool, str]:
    """判断一个值是否是时间，返回 (是否时间, 标准化标签)。
    
    支持的格式:
    - 2024 (年份数字)              → "2024"
    - 45292 (Excel日期序列号)      → "2024-01"
    - "2024-01", "2024年1月"      → "2024-01"
    - "2024Q1", "Q1-2024"         → "2024-Q1"
    - "FY24", "FY2024"            → "2024"
    - "Jan-24", "1月"              → "2024-01"
    """
    if v is None:
        return False, ""
    
    # 1. Excel 日期序列号
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if v != int(v):
            return False, ""
        if 36526 <= v <= 73050:  # 2000-01-01 to 2099-12-31
            return True, _excel_serial_to_label(v)
        if 2000 <= v <= 2100:
            return True, str(int(v))
        return False, ""
    
    # 2. 字符串格式
    if not isinstance(v, str):
        return False, ""
    
    s = v.strip()
    if not s:
        return False, ""
    
    # 纯年份: "2024"
    if re.match(r"^20\d{2}$", s):
        return True, s
    
    # 年月: "2024-01", "2024/01", "2024.01"
    m = re.match(r"^(20\d{2})[-/\.](\d{1,2})$", s)
    if m:
        return True, f"{m.group(1)}-{int(m.group(2)):02d}"
    
    # 中文年月: "2024年1月", "2024年01月"
    m = re.match(r"^(20\d{2})年(\d{1,2})月$", s)
    if m:
        return True, f"{m.group(1)}-{int(m.group(2)):02d}"
    
    # 季度: "2024Q1", "2024q1"
    m = re.match(r"^(20\d{2})[Qq](\d)$", s)
    if m:
        return True, f"{m.group(1)}-Q{m.group(2)}"
    
    # 季度: "Q1-2024", "Q1/2024"
    m = re.match(r"^[Qq](\d)[-/](20\d{2})$", s)
    if m:
        return True, f"{m.group(2)}-Q{m.group(1)}"
    
    # 财年: "FY2024", "fy2024"
    m = re.match(r"^[Ff][Yy](20\d{2})$", s)
    if m:
        return True, m.group(1)
    
    # 财年缩写: "FY24" → 假设 2000s
    m = re.match(r"^[Ff][Yy](\d{2})$", s)
    if m:
        year = int(m.group(1))
        full_year = 2000 + year if year < 100 else year
        return True, str(full_year)
    
    # 英文月年: "Jan-24", "Feb-24"
    month_map = {
        "jan": "01", "feb": "02", "mar": "03", "apr": "04",
        "may": "05", "jun": "06", "jul": "07", "aug": "08",
        "sep": "09", "oct": "10", "nov": "11", "dec": "12",
    }
    m = re.match(r"^([A-Za-z]{3})[-](\d{2})$", s)
    if m:
        month_abbr = m.group(1).lower()
        year_suffix = int(m.group(2))
        if month_abbr in month_map:
            full_year = 2000 + year_suffix if year_suffix < 100 else year_suffix
            return True, f"{full_year}-{month_map[month_abbr]}"
    
    return False, ""


def _is_excel_date_serial(v) -> bool:
    """兼容旧接口: 判断是否为 Excel 日期序列号。"""
    is_time, _ = _is_time_value(v)
    return is_time and isinstance(v, (int, float)) and not isinstance(v, bool) and 36526 <= v <= 73050


def _is_year_value(v) -> bool:
    """兼容旧接口: 判断是否为年份数字。"""
    is_time, label = _is_time_value(v)
    return is_time and isinstance(v, (int, float)) and not isinstance(v, bool) and 2000 <= v <= 2100


def _looks_like_sequence(v) -> bool:
    if isinstance(v, (int, float)):
        return 0 < v < 1000
    if isinstance(v, str):
        return bool(re.match(r"^\d+(\.\d+)*$", v.strip()))
    return False


# --- Table structure ---

class ColRole:
    CATEGORY = "category"
    SEQUENCE = "sequence"
    NAME = "name"
    TOTAL = "total"
    UNIT = "unit"
    TIME_SERIES = "time_series"
    NOTES = "notes"
    FORMULA_DESC = "formula_desc"
    UNKNOWN = "unknown"


class TableInfo:
    """Detected table within a sheet."""

    def __init__(self, sheet: str, header_row: int, data_start: int, data_end: int,
                 title: Optional[str] = None,
                 start_col: Optional[str] = None, end_col: Optional[str] = None):
        self.sheet = sheet
        self.header_row = header_row
        self.data_start = data_start
        self.data_end = data_end
        self.title = title
        self.start_col = start_col
        self.end_col = end_col
        self.col_roles: dict[str, str] = {}
        self.col_labels: dict[str, str] = {}
        self.time_period_labels: dict[str, str] = {}
        self.is_double_header: bool = False  # True if header spans 2 rows

    def name_col(self) -> Optional[str]:
        for col, role in self.col_roles.items():
            if role == ColRole.NAME:
                return col
        return None

    def total_col(self) -> Optional[str]:
        for col, role in self.col_roles.items():
            if role == ColRole.TOTAL:
                return col
        return None

    def unit_col(self) -> Optional[str]:
        for col, role in self.col_roles.items():
            if role == ColRole.UNIT:
                return col
        return None

    def category_col(self) -> Optional[str]:
        for col, role in self.col_roles.items():
            if role == ColRole.CATEGORY:
                return col
        return None

    def sequence_col(self) -> Optional[str]:
        for col, role in self.col_roles.items():
            if role == ColRole.SEQUENCE:
                return col
        return None

    def time_series_cols(self) -> list[str]:
        return [col for col, role in self.col_roles.items() if role == ColRole.TIME_SERIES]

# --- Merged cell helpers ---

def _build_merge_groups(cell_list: list) -> dict:
    """Build merge group info from CellData list.

    Uses merge_end_row/merge_end_col fields on top-left cells.
    Returns: {parent_id: {"rows": set, "cols": set, "cells": set}}
    """
    groups: dict = {}
    for cd in cell_list:
        if cd.merge_end_row is not None and cd.merge_end_col is not None:
            parent_id = cd.id
            groups[parent_id] = {
                "rows": set(range(cd.row, cd.merge_end_row + 1)),
                "cols": set(),
                "cells": set(),
            }
            # Fill in cols and cells
            for c_idx in range(
                column_index_from_string(cd.col),
                column_index_from_string(cd.merge_end_col) + 1,
            ):
                col_letter = get_column_letter(c_idx)
                groups[parent_id]["cols"].add(col_letter)
                for r in range(cd.row, cd.merge_end_row + 1):
                    groups[parent_id]["cells"].add((r, col_letter))
    return groups


def _find_vertical_anchors(
    merge_groups: dict,
    rows: dict[int, dict[str, object]],
    sheet_name: str,
) -> list[dict]:
    """Find vertical merged cells that serve as L-table anchors.

    Criteria: >= 3 rows tall, <= 2 cols wide, non-empty top-left value.
    Returns sorted list of {start_row, start_col, end_row, end_col, parent_id}.
    """
    anchors = []
    for parent_id, info in merge_groups.items():
        rows_set = info["rows"]
        cols_set = info["cols"]
        if len(rows_set) < 3:
            continue
        if len(cols_set) > 2:
            continue
        # The parent cell is the top-left of the merge group
        # Parse parent_id: "sheet_row_col" -> extract row and col
        parts = parent_id.split("_")
        if len(parts) < 3:
            continue
        try:
            anchor_row = int(parts[-2])
            anchor_col = parts[-1]
        except (ValueError, IndexError):
            continue
        if parent_id.startswith(sheet_name + "_"):
            if anchor_row in rows and anchor_col in rows[anchor_row]:
                anchors.append({
                    "start_row": min(rows_set),
                    "start_col": min(cols_set, key=column_index_from_string),
                    "end_row": max(rows_set),
                    "end_col": max(cols_set, key=column_index_from_string),
                    "parent_id": parent_id,
                })
    anchors.sort(key=lambda a: (a["start_row"], column_index_from_string(a["start_col"])))
    return anchors


def _col_offset(col: str, delta: int) -> str | None:
    idx = column_index_from_string(col) + delta
    if idx < 1:
        return None
    return get_column_letter(idx)


def _looks_like_table_header(row_data: dict) -> bool:
    """Check if a row looks like a column header row (not a data row).

    After a gap, this distinguishes a new table's header row from
    the resumption of an L-table after a section break.

    A header row: mostly/all text labels, no numeric values, and
    at least two known header keywords to avoid false positives.
    """
    if not row_data:
        return False
    values = [v for v in row_data.values() if v is not None]
    if not values:
        return False
    str_vals = [v for v in values if isinstance(v, str) and v.strip()]
    num_vals = [v for v in values if isinstance(v, (int, float)) and not isinstance(v, bool)]
    if num_vals:
        return False
    if len(str_vals) < 2:
        return False
    matches = sum(1 for v in str_vals if any(kw in v for kw in _HEADER_KEYWORDS))
    return matches >= 2


def _is_in_recognized_range(
    row: int, col: str, recognized_ranges: list[tuple[int, str, int, str]],
) -> bool:
    for sr, sc, er, ec in recognized_ranges:
        if sr <= row <= er and column_index_from_string(sc) <= column_index_from_string(col) <= column_index_from_string(ec):
            return True
    return False


def _connected_column_groups(
    row_data: dict[str, object], min_col_idx: int, max_gap: int = 5,
) -> list[list[str]]:
    """Group non-empty columns into connected clusters by column index gap.

    Two columns with gap ≤ max_gap empty columns between them are in the same group.
    Only considers columns at or to the right of min_col_idx.
    """
    cols = sorted(
        [c for c in row_data if row_data[c] is not None
         and column_index_from_string(c) >= min_col_idx],
        key=column_index_from_string,
    )
    if not cols:
        return []
    groups: list[list[str]] = []
    current = [cols[0]]
    for i in range(1, len(cols)):
        gap = column_index_from_string(cols[i]) - column_index_from_string(cols[i - 1]) - 1
        if gap <= max_gap:
            current.append(cols[i])
        else:
            groups.append(current)
            current = [cols[i]]
    groups.append(current)
    return groups


# --- Main detection function ---

def detect_tables(
    sheet_name: str,
    rows: dict[int, dict[str, object]],
    cell_list: list | None = None,
) -> list[TableInfo]:
    """Detect logical tables within a sheet using spatial analysis.

    Algorithm (adapted from excel_table_extractor.py):
    1. Build merge groups from CellData -> find vertical anchors
    2. Row-scan-expand L-tables from each vertical anchor (L-tables first)
    3. BFS-expand rectangular tables from remaining unvisited cells
    4. Merge adjacent tables with small column gaps
    5. Filter trivial tables (<=1 row or <=1 cell)

    Args:
        sheet_name: Name of the sheet.
        rows: Dict mapping row number to dict of {col_letter: cell_value}.
        cell_list: Optional list of CellData for merge info.

    Returns:
        List of TableInfo objects.
    """
    if not rows:
        return []

    merge_groups = _build_merge_groups(cell_list) if cell_list else {}
    anchors = _find_vertical_anchors(merge_groups, rows, sheet_name)

    max_row = max(rows.keys())
    all_cols = set()
    for row_data in rows.values():
        all_cols.update(row_data.keys())
    max_col = max(all_cols, key=column_index_from_string) if all_cols else "A"

    visited: set[tuple[int, str]] = set()
    recognized_ranges: list[tuple[int, str, int, str]] = []
    tables: list[TableInfo] = []

    # --- Phase 1: Identify anchor territory ---
    anchor_territories: list[tuple[int, int, str, str]] = []  # (start_row, end_row, start_col, end_col)
    for a in anchors:
        territory_end_col = _col_offset(a["start_col"], 100)
        if territory_end_col is None:
            territory_end_col = max_col
        anchor_territories.append((max(1, a["start_row"] - 1), a["end_row"], a["start_col"], territory_end_col))

    # --- Phase 2: L-table expansion from each anchor (FIRST, so
    # L-tables get priority over rectangular BFS expansion) ---
    for a in anchors:
        if (a["start_row"], a["start_col"]) in visited:
            continue

        # Pre-scan: detect disconnected sub-table immediately above the anchor.
        # If row ar-1 has data far from the anchor column (>4 gap) AND looks
        # like a table header, BFS-expand it as a separate table first so the
        # L-table won't absorb its cells (Problem 3: J-N table rows 182-187).
        ar, ac = a["start_row"], a["start_col"]
        if ar > 1:
            pre_row = rows.get(ar - 1, {})
            pre_cols = sorted(
                [c for c in pre_row if pre_row[c] is not None],
                key=column_index_from_string,
            )
            if pre_cols:
                min_pre_col_idx = column_index_from_string(pre_cols[0])
                anchor_col_idx = column_index_from_string(ac)
                if min_pre_col_idx - anchor_col_idx > 4 and _looks_like_table_header(pre_row):
                    sub = _bfs_expand_rect_table(
                        ar - 1, pre_cols[0], rows, visited, max_row, max_col, recognized_ranges,
                    )
                    if sub:
                        recognized_ranges.append((
                            sub["start_row"], sub["start_col"],
                            sub["end_row"], sub["end_col"],
                        ))
                        tables.append(sub)

        table_info = _expand_l_table(
            a, rows, visited, max_row, max_col, recognized_ranges,
            other_anchors=[
                a2 for a2 in anchors
                if (a2["start_row"], a2["start_col"]) not in visited
                and a2["start_row"] > a["start_row"]
            ],
        )
        if table_info is None:
            continue

        # Find overlapping recognized ranges -> excluded_ranges
        excluded = []
        sr, sc, er, ec = table_info["start_row"], table_info["start_col"], table_info["end_row"], table_info["end_col"]
        for rr in recognized_ranges:
            rsr, rsc, rer, rec = rr
            if (rsr <= er and rer >= sr and
                column_index_from_string(rsc) <= column_index_from_string(ec) and
                column_index_from_string(rec) >= column_index_from_string(sc)):
                excluded.append(rr)
        table_info["excluded_ranges"] = excluded

        tables.append(table_info)
        recognized_ranges.append((sr, sc, er, ec))

        # Mark inner anchors as visited
        for a2 in anchors:
            if sr <= a2["start_row"] <= er and column_index_from_string(sc) <= column_index_from_string(a2["start_col"]) <= column_index_from_string(ec):
                visited.add((a2["start_row"], a2["start_col"]))

    # --- Phase 3: BFS-expand rectangular tables from remaining unvisited cells ---
    sorted_positions = []
    for r in sorted(rows.keys()):
        for c in sorted(rows[r].keys(), key=column_index_from_string):
            sorted_positions.append((r, c))

    for r, c in sorted_positions:
        if (r, c) in visited:
            continue

        # Skip L-anchor top-left cells
        is_anchor_top = False
        for a in anchors:
            if a["start_row"] == r and a["start_col"] == c:
                is_anchor_top = True
                break
        if is_anchor_top:
            continue

        # Skip cells in anchor territory
        in_territory = False
        for tr, ter, tc, tec in anchor_territories:
            if tr <= r <= ter and column_index_from_string(tc) <= column_index_from_string(c):
                in_territory = True
                break
        if in_territory:
            continue

        table_info = _bfs_expand_rect_table(
            r, c, rows, visited, max_row, max_col, recognized_ranges,
        )
        if table_info:
            tables.append(table_info)
            recognized_ranges.append((
                table_info["start_row"], table_info["start_col"],
                table_info["end_row"], table_info["end_col"],
            ))

    # --- Phase 4: Merge adjacent tables ---
    tables = _merge_adjacent_tables(tables, rows, max_row, max_col)

    # --- Phase 4.5: Merge overlapping tables (same row range) ---
    tables = _merge_overlapping_tables(tables, rows)

    # Sort
    tables.sort(key=lambda t: (t["start_row"], column_index_from_string(t["start_col"])))

    # --- Phase 5: Convert to TableInfo and filter ---
    result: list[TableInfo] = []
    for raw in tables:
        tbl = _raw_to_table_info(sheet_name, raw, rows)
        if tbl is not None:
            result.append(tbl)

    # --- Phase 5.5: Merge overlapping TableInfo objects ---
    result = _merge_overlapping_table_infos(result, rows)

    # --- Phase 6: Per-sheet numbering and title formatting ---
    for idx, tbl in enumerate(result, 1):
        sheet_short = sheet_name.strip()
        if tbl.title:
            tbl.title = f"{sheet_short}-{idx}-{tbl.title}"
        else:
            tbl.title = f"{sheet_short}-{idx}"

    return result


# --- Rectangular table expansion (BFS) ---

def _bfs_expand_rect_table(
    start_row: int, start_col: str,
    rows: dict[int, dict[str, object]],
    visited: set,
    max_row: int, max_col: str,
    recognized_ranges: list,
) -> dict | None:
    """BFS from a starting cell to find a rectangular connected block."""
    table_cells: set[tuple[int, str]] = set()
    queue = [(start_row, start_col)]

    # Track which rows we've already added to the queue to avoid
    # re-scanning header checks for the same row multiple times
    header_rows_seen: set[int] = set()

    # Determine if this BFS started from a pure title row
    start_row_data = rows.get(start_row, {})
    start_row_vals = [v for v in start_row_data.values() if v is not None]
    start_row_strs = [v for v in start_row_vals if isinstance(v, str) and v.strip()]
    start_row_nums = [v for v in start_row_vals if isinstance(v, (int, float)) and not isinstance(v, bool)]
    is_title_start = (len(start_row_vals) <= 2 and len(start_row_strs) >= 1
                      and len(start_row_nums) == 0 and len(header_rows_seen) == 0)

    while queue:
        r, c = queue.pop(0)
        if (r, c) in visited:
            continue
        visited.add((r, c))

        # Check if in recognized range
        if _is_in_recognized_range(r, c, recognized_ranges):
            continue

        if r in rows and c in rows[r] and rows[r][c] is not None:
            table_cells.add((r, c))
        else:
            continue

        # Explore neighbors
        for nr, nc in [(r - 1, c), (r + 1, c), (r, _col_offset(c, -1)), (r, _col_offset(c, 1))]:
            if nc is None:
                continue
            if nr < 1 or nr > max_row:
                continue
            if column_index_from_string(nc) > column_index_from_string(max_col):
                continue
            if (nr, nc) in visited:
                continue

            # Prevent BFS from absorbing proper header rows of subsequent tables.
            # Only apply this check when the BFS started from a pure title row
            # (e.g., row 243 "折旧摊销表-全投资..." in 表2). This prevents the
            # title-only BFS from absorbing the header row below it.
            if is_title_start and nr not in header_rows_seen and nr != start_row:
                header_rows_seen.add(nr)
                row_data = rows.get(nr, {})
                header_kw_count = sum(
                    1 for v in row_data.values()
                    if isinstance(v, str) and v.strip() and any(
                        kw in v for kw in _SEQ_KEYWORDS | _NAME_KEYWORDS | _CATEGORY_KEYWORDS
                    )
                )
                if header_kw_count >= 2:
                    # This row is a proper header — skip all cells in this row
                    for skip_c in row_data:
                        visited.add((nr, skip_c))
                    continue

            queue.append((nr, nc))

    if not table_cells:
        return None

    rs = [p[0] for p in table_cells]
    cs = [column_index_from_string(p[1]) for p in table_cells]
    return {
        "start_row": min(rs), "end_row": max(rs),
        "start_col": get_column_letter(min(cs)), "end_col": get_column_letter(max(cs)),
        "cells": table_cells,
    }


# --- L-table expansion (row-scanning) ---

def _expand_l_table(
    anchor: dict,
    rows: dict[int, dict[str, object]],
    visited: set,
    max_row: int, max_col: str,
    recognized_ranges: list,
    other_anchors: list[dict] | None = None,
) -> dict | None:
    """Expand an L-shaped table from a vertical anchor.

    Key difference from BFS: scans entire rows until an empty row,
    tolerating sparse columns within the table.

    If other_anchors is provided, stops expansion when encountering
    the start_row of another anchor (to split adjacent L-tables).
    """
    ar, ac, aer, aec = anchor["start_row"], anchor["start_col"], anchor["end_row"], anchor["end_col"]

    table_cells: set[tuple[int, str]] = set()

    # Add anchor region cells
    for r in range(ar, aer + 1):
        for c_idx in range(column_index_from_string(ac), column_index_from_string(aec) + 1):
            c = get_column_letter(c_idx)
            table_cells.add((r, c))
            visited.add((r, c))

    # Find max data column: scan anchor row + next 50 rows for rightmost data
    max_data_col_idx = column_index_from_string(aec)
    for r in range(ar, min(ar + 51, max_row + 1)):
        if r not in rows:
            continue
        for c in sorted(rows[r].keys(), key=column_index_from_string):
            c_idx = column_index_from_string(c)
            if c_idx > column_index_from_string(aec) and rows[r][c] is not None:
                max_data_col_idx = max(max_data_col_idx, c_idx)

    # Row-scan downward with gap tolerance: skip ≤2 consecutive empty rows,
    # break at 3+ empty rows. This allows L-tables to span section breaks
    # that are separated by 1-2 empty rows.
    # Start one row above anchor to capture header row (e.g. row 3 before
    # anchor at row 4), but don't go below row 1.
    current_row = ar
    if ar > 1:
        pre_row = rows.get(ar - 1, {})
        pre_cols = [column_index_from_string(c) for c in pre_row if pre_row[c] is not None]
        if pre_cols:
            min_pre_col = min(pre_cols)
            anchor_col_idx = column_index_from_string(ac)
            if min_pre_col - anchor_col_idx <= 4:
                # Adjacent — extend upward to capture header/title
                current_row = ar - 1
            # else: disconnected — leave for BFS, start from anchor row

    consecutive_empty = 0
    while current_row <= max_row:
        row_has_data = False
        for c_idx in range(column_index_from_string(ac), max_data_col_idx + 1):
            c = get_column_letter(c_idx)
            if _is_in_recognized_range(current_row, c, recognized_ranges):
                row_has_data = True
                break
            if current_row in rows and c in rows[current_row] and rows[current_row][c] is not None:
                row_has_data = True
                break

        if not row_has_data:
            if ar <= current_row <= aer:
                # Empty rows within the anchor range are always OK
                for c_idx in range(column_index_from_string(ac), max_data_col_idx + 1):
                    visited.add((current_row, get_column_letter(c_idx)))
                current_row += 1
                consecutive_empty = 0
                continue

            consecutive_empty += 1
            if consecutive_empty >= 3:
                break
            # Mark this empty row's cells as visited
            for c_idx in range(column_index_from_string(ac), max_data_col_idx + 1):
                visited.add((current_row, get_column_letter(c_idx)))
            current_row += 1
            continue

        # Data resumed after a gap — check for new table or orphan note
        if consecutive_empty > 0:
            row_data = rows.get(current_row, {})
            if _looks_like_table_header(row_data):
                break
            # Check if this row is the start of another L-table anchor
            is_new_anchor = False
            if other_anchors:
                for oa in other_anchors:
                    if oa["start_row"] == current_row:
                        is_new_anchor = True
                        break
            if is_new_anchor:
                break
            # Single-cell text row after a gap is likely an orphan note,
            # not part of the same table
            non_empty = [v for v in row_data.values() if v is not None]
            if len(non_empty) == 1 and all(isinstance(v, str) and not v.strip().isdigit() for v in non_empty):
                break

        consecutive_empty = 0

        # Add cells in this row — only from the column group connected to
        # the anchor column. Disconnected groups (gap > 5 empty cols) are
        # left unvisited for BFS Phase 3 to pick up as separate tables.
        row_data = rows.get(current_row, {})
        groups = _connected_column_groups(row_data, column_index_from_string(ac), max_gap=5)
        anchor_group: list[str] = []
        for g in groups:
            if ac in g:
                anchor_group = g
                break
        if not anchor_group and groups:
            # No group contains anchor — pick the leftmost group closest to anchor
            anchor_group = min(groups, key=lambda g: column_index_from_string(g[0]))

        anchor_col_set = set(anchor_group)
        for c_idx in range(column_index_from_string(ac), max_data_col_idx + 1):
            c = get_column_letter(c_idx)
            if c not in anchor_col_set:
                continue
            if _is_in_recognized_range(current_row, c, recognized_ranges):
                continue
            if (current_row, c) in visited:
                continue
            if current_row in rows and c in rows[current_row] and rows[current_row][c] is not None:
                table_cells.add((current_row, c))
            visited.add((current_row, c))

        current_row += 1

    if not table_cells:
        return None

    rs = [p[0] for p in table_cells]
    cs = [column_index_from_string(p[1]) for p in table_cells]

    # Use anchor's top-left cell value as fallback title (e.g. "季度还款")
    title_text = None
    anchor_val = rows.get(ar, {}).get(ac)
    if anchor_val is not None and isinstance(anchor_val, str) and anchor_val.strip():
        title_text = str(anchor_val).strip()

    return {
        "start_row": min(rs), "end_row": max(rs),
        "start_col": get_column_letter(min(cs)), "end_col": get_column_letter(max(cs)),
        "cells": table_cells,
        "title_text": title_text,
    }


# --- Merge adjacent tables ---

def _merge_adjacent_tables(
    tables: list[dict],
    rows: dict[int, dict[str, object]],
    max_row: int, max_col: str,
) -> list[dict]:
    """Merge tables on similar rows with small column gaps and bridging content."""
    if not tables:
        return tables

    merged = []
    used = set()
    for i, t1 in enumerate(tables):
        if i in used:
            continue
        current = dict(t1)
        for j, t2 in enumerate(tables):
            if j <= i or j in used:
                continue
            col_gap = column_index_from_string(t2["start_col"]) - column_index_from_string(current["end_col"]) - 1
            row_overlap = (current["start_row"] <= t2["end_row"] and t2["start_row"] <= current["end_row"])

            if row_overlap and 0 < col_gap <= 3:
                # Check for bridging content
                title_row = min(current["start_row"], t2["start_row"]) - 1
                has_bridge = False
                for check_row in [title_row, min(current["start_row"], t2["start_row"])]:
                    if 1 <= check_row <= max_row and check_row in rows:
                        for c_idx in range(
                            column_index_from_string(current["start_col"]),
                            max(column_index_from_string(current["end_col"]), column_index_from_string(t2["end_col"])) + 1,
                        ):
                            c = get_column_letter(c_idx)
                            if c in rows[check_row] and rows[check_row][c] is not None:
                                has_bridge = True
                                break
                    if has_bridge:
                        break

                if has_bridge:
                    # Preserve current's excluded_ranges before recreating dict
                    existing_excluded = list(current.get("excluded_ranges", []))
                    existing_title = current.get("title_text")
                    new_cells = current.get("cells", set()) | t2.get("cells", set())
                    new_sr = min(current["start_row"], t2["start_row"])
                    new_er = max(current["end_row"], t2["end_row"])
                    new_sc = get_column_letter(min(
                        column_index_from_string(current["start_col"]),
                        column_index_from_string(t2["start_col"]),
                    ))
                    new_ec = get_column_letter(max(
                        column_index_from_string(current["end_col"]),
                        column_index_from_string(t2["end_col"]),
                    ))
                    current = {
                        "start_row": new_sr, "end_row": new_er,
                        "start_col": new_sc, "end_col": new_ec,
                        "cells": new_cells,
                    }
                    # Carry forward title from current or adopt from t2
                    if existing_title:
                        current["title_text"] = existing_title
                    elif "title_text" in t2:
                        current["title_text"] = t2["title_text"]
                    # Merge excluded_ranges from both tables
                    current["excluded_ranges"] = existing_excluded + list(t2.get("excluded_ranges", []))
                    used.add(j)

        merged.append(current)
        used.add(i)

    return merged


def _merge_overlapping_tables(
    tables: list[dict],
    rows: dict[int, dict[str, object]],
) -> list[dict]:
    """Merge tables with identical or nearly identical row ranges.

    When BFS expansion finds overlapping connected components (e.g., two tables
    with the same start_row and end_row but different column groups), merge them
    into a single table.

    Criteria for merging:
    - Same start_row and end_row (or within 1 row)
    - Column ranges overlap or are adjacent
    """
    if len(tables) <= 1:
        return tables

    merged = []
    used = set()

    for i, t1 in enumerate(tables):
        if i in used:
            continue

        current = dict(t1)
        current_cells = set(t1.get("cells", set()))

        for j, t2 in enumerate(tables):
            if j <= i or j in used:
                continue

            # Check if row ranges are identical or within 1 row
            row_diff = abs(t1["start_row"] - t2["start_row"]) + abs(t1["end_row"] - t2["end_row"])
            if row_diff > 2:
                continue

            # Check if column ranges overlap or are adjacent
            t1_sc = column_index_from_string(t1["start_col"])
            t1_ec = column_index_from_string(t1["end_col"])
            t2_sc = column_index_from_string(t2["start_col"])
            t2_ec = column_index_from_string(t2["end_col"])

            col_overlap = t1_ec >= t2_sc and t2_ec >= t1_sc
            if not col_overlap:
                continue

            # Merge: combine cells if available
            t2_cells = set(t2.get("cells", set()))
            if current_cells or t2_cells:
                current_cells = current_cells | t2_cells
                current["cells"] = current_cells
                all_cols = {c for r, c in current_cells}
                if all_cols:
                    current["start_col"] = min(all_cols, key=column_index_from_string)
                    current["end_col"] = max(all_cols, key=column_index_from_string)
                else:
                    current["start_col"] = min(t1["start_col"], t2["start_col"], key=column_index_from_string)
                    current["end_col"] = max(t1["end_col"], t2["end_col"], key=column_index_from_string)
            else:
                current["start_col"] = min(t1["start_col"], t2["start_col"], key=column_index_from_string)
                current["end_col"] = max(t1["end_col"], t2["end_col"], key=column_index_from_string)

            current["start_row"] = min(current["start_row"], t2["start_row"])
            current["end_row"] = max(current["end_row"], t2["end_row"])

            # Carry forward title from current or t2
            if "title_text" not in current and "title_text" in t2:
                current["title_text"] = t2["title_text"]

            used.add(j)

        merged.append(current)
        used.add(i)

    return merged


def _merge_overlapping_table_infos(
    tables: list[TableInfo],
    rows: dict[int, dict[str, object]],
) -> list[TableInfo]:
    """Merge TableInfo objects with identical data_row_range.

    After _raw_to_table_info normalizes header/data rows, some tables may end
    up with identical data_row_range but were detected from different raw
    tables. Merge them to avoid duplicates.
    """
    if len(tables) <= 1:
        return tables

    merged = []
    used = set()

    for i, t1 in enumerate(tables):
        if i in used:
            continue

        current = t1
        for j, t2 in enumerate(tables):
            if j <= i or j in used:
                continue

            # Check if data row range is identical
            if (current.data_start == t2.data_start and
                current.data_end == t2.data_end and
                current.header_row == t2.header_row and
                current.sheet == t2.sheet):
                # Merge: combine col_roles and time_period_labels
                for col, role in t2.col_roles.items():
                    if col not in current.col_roles:
                        current.col_roles[col] = role
                    elif role == ColRole.TIME_SERIES:
                        current.col_roles[col] = role

                for col, label in t2.time_period_labels.items():
                    if col not in current.time_period_labels:
                        current.time_period_labels[col] = label

                # Merge col_labels
                for col, label in t2.col_labels.items():
                    if col not in current.col_labels:
                        current.col_labels[col] = label

                used.add(j)

        merged.append(current)
        used.add(i)

    return merged


# --- Conversion to TableInfo ---

def _raw_to_table_info(
    sheet_name: str, raw: dict,
    rows: dict[int, dict[str, object]],
) -> Optional[TableInfo]:
    """Convert a raw table dict to a TableInfo, filtering trivial tables."""
    cells = raw.get("cells", set())
    if len(cells) <= 1:
        return None

    rows_in = sorted({r for r, c in cells})
    if len(rows_in) <= 1:
        return None

    header_row = rows_in[0]
    data_end = rows_in[-1]

    # ── P1-1 fix: Detect header rows above the detected first row ────────────
    # BFS/L-table expansion may start from the first data row but miss merged
    # header rows above. Check 1-2 rows above for header-like content.
    _HEADER_KW = _SEQ_KEYWORDS | _NAME_KEYWORDS | _CATEGORY_KEYWORDS
    for candidate_row in range(header_row - 1, max(header_row - 3, 0), -1):
        candidate_data = rows.get(candidate_row, {})
        candidate_vals = [v for v in candidate_data.values() if v is not None]
        if not candidate_vals:
            continue
        # Check if this row has header keywords
        header_kw_count = sum(
            1 for v in candidate_vals
            if isinstance(v, str) and v.strip() and any(kw in v for kw in _HEADER_KW)
        )
        if header_kw_count >= 2:
            # This row looks like a proper header above the detected table
            header_row = candidate_row
            break

    # Detect if first row is a title (1-2 text cells, no header keywords)
    # rather than a real data/header row. E.g. "工程分年度投资输入表"
    first_row_data = rows.get(header_row, {})
    first_vals = [v for v in first_row_data.values() if v is not None]
    first_str_vals = [v for v in first_vals if isinstance(v, str) and v.strip()]
    first_num_vals = [v for v in first_vals
                     if isinstance(v, (int, float)) and not isinstance(v, bool)
                     and not _is_excel_date_serial(v)]
    first_has_header_kw = any(
        any(kw in (str(v) if v else "") for kw in _SEQ_KEYWORDS | _NAME_KEYWORDS | _CATEGORY_KEYWORDS)
        for v in first_vals
    )
    if (len(first_vals) <= 2 and len(first_str_vals) >= 1
            and len(first_num_vals) == 0 and not first_has_header_kw
            and len(rows_in) >= 3):
        # First row is a title — promote it and advance header
        title = first_str_vals[0]
        header_row = rows_in[1]
        header_data = rows.get(header_row, {})
    else:
        title = None  # Will be determined by title search or fallback below
        header_data = first_row_data

    # Determine whether this is a header row or data-first row.
    # L-tables may have no column header — e.g. anchor text + field-value pairs.
    num_count = sum(1 for v in header_data.values()
                    if isinstance(v, (int, float)) and not isinstance(v, bool)
                    and not _is_excel_date_serial(v))
    date_count = sum(1 for v in header_data.values() if _is_excel_date_serial(v))
    header_labels = sum(1 for v in header_data.values()
                       if isinstance(v, str) and v.strip()
                       and any(kw in v for kw in _SEQ_KEYWORDS | _NAME_KEYWORDS | _CATEGORY_KEYWORDS))
    has_proper_header = (
        header_labels >= 3
        or (len(header_data) >= 5 and num_count > 0)
        or (date_count >= 3 and len(header_data) >= 3)
    )
    if has_proper_header:
        data_start = header_row + 1
    elif num_count > 0 and len(header_data) > 2:
        # Numeric header row (unusual but possible)
        data_start = header_row
    else:
        # No proper header — data starts at header_row (L-table label-value pattern)
        data_start = header_row

    if data_start > data_end:
        return None

    # Require at least 2 rows of actual data (filter single-row orphans)
    if data_end - data_start + 1 < 2:
        return None

    # Title search (if not already found from first-row promotion)
    if not title:
        for title_row_num in range(header_row - 1, max(header_row - 4, 0), -1):
            candidate = rows.get(title_row_num, {})
            vals = [v for v in candidate.values() if v is not None]
            str_vals = [v for v in vals if isinstance(v, str) and v.strip()]
            num_vals = [v for v in vals if isinstance(v, (int, float)) and not isinstance(v, bool)]
            if len(vals) >= 1 and len(str_vals) >= 1 and len(num_vals) == 0:
                title = str_vals[0]
                break

    # Fall back to anchor text if no title found above header
    if not title:
        title = raw.get("title_text")

    start_col = raw.get("start_col")
    end_col = raw.get("end_col")
    tbl = TableInfo(sheet_name, header_row, data_start, data_end,
                    title=title, start_col=start_col, end_col=end_col)
    _classify_columns(tbl, rows)
    return tbl


# --- Column classification ---

def _classify_columns(tbl: TableInfo, rows: dict[int, dict[str, object]]) -> None:
    header_row = rows.get(tbl.header_row, {})

    data_cols: set[str] = set()
    for rnum in range(tbl.data_start, tbl.data_end + 1):
        if rnum in rows:
            data_cols.update(rows[rnum].keys())
    data_cols.update(header_row.keys())

    # Restrict to the table's column bounds if known
    if tbl.start_col and tbl.end_col:
        sc_idx = column_index_from_string(tbl.start_col)
        ec_idx = column_index_from_string(tbl.end_col)
        data_cols = {c for c in data_cols
                     if sc_idx <= column_index_from_string(c) <= ec_idx}

    # Detect double-header: if the current header_row has mostly numbers
    # (like month numbers 1-12) and the next row has date serials,
    # use the next row as the time reference.
    next_row_data = rows.get(tbl.header_row + 1, {})
    has_time_in_next = sum(1 for v in next_row_data.values() if _is_time_value(v)[0])
    use_time_row = tbl.header_row + 1 if has_time_in_next > 2 else None

    # Also detect double-header: header keywords in header_row + sequence numbers in next row
    # Only if time_period_labels is empty (meaning time values were not found in header_row)
    # and the next row has sequence numbers (1, 2, 3...) across time columns
    if not use_time_row and next_row_data and not tbl.time_period_labels:
        header_kw_count = sum(
            1 for v in header_row.values()
            if isinstance(v, str) and v.strip() and any(kw in v for kw in _SEQ_KEYWORDS | _NAME_KEYWORDS | _CATEGORY_KEYWORDS)
        )
        # Count sequence numbers in the next row that appear in columns where header_row has values
        seq_in_next = sum(
            1 for c, v in next_row_data.items()
            if c in header_row and isinstance(v, (int, float)) and not isinstance(v, bool) and 1 <= v <= 100
        )
        # Only set double-header if the header row has proper header keywords AND
        # the next row has sequence numbers in matching columns
        if header_kw_count >= 2 and seq_in_next >= 4:
            tbl.is_double_header = True

    for col in sorted(data_cols, key=lambda c: column_index_from_string(c)):
        label = header_row.get(col)
        label_str = str(label).strip() if label is not None else ""

        # For double-header tables, get time value from the next row
        time_label = None
        if use_time_row and col in next_row_data:
            is_time, period_label = _is_time_value(next_row_data[col])
            if is_time:
                time_label = period_label

        role = _role_from_label(label_str)

        if role == ColRole.UNKNOWN and label is not None:
            is_time, period_label = _is_time_value(label)
            if is_time:
                role = ColRole.TIME_SERIES
                tbl.time_period_labels[col] = period_label

        # Apply time from double-header row
        if time_label is not None and role == ColRole.UNKNOWN:
            role = ColRole.TIME_SERIES
            tbl.time_period_labels[col] = time_label

        if role == ColRole.UNKNOWN:
            role = _role_from_data(col, tbl, rows)

        tbl.col_roles[col] = role
        tbl.col_labels[col] = label_str


def _role_from_label(label: str) -> str:
    if not label:
        return ColRole.UNKNOWN
    for kw in _CATEGORY_KEYWORDS:
        if kw in label:
            return ColRole.CATEGORY
    for kw in _SEQ_KEYWORDS:
        if kw in label:
            return ColRole.SEQUENCE
    for kw in _NAME_KEYWORDS:
        if kw in label:
            return ColRole.NAME
    for kw in _TOTAL_KEYWORDS:
        if kw in label:
            return ColRole.TOTAL
    for kw in _UNIT_KEYWORDS:
        if kw in label:
            return ColRole.UNIT
    for kw in _NOTES_KEYWORDS:
        if kw in label:
            return ColRole.NOTES
    return ColRole.UNKNOWN


def _role_from_data(col: str, tbl: TableInfo, rows: dict[int, dict[str, object]]) -> str:
    sample = []
    for rnum in range(tbl.data_start, min(tbl.data_start + 10, tbl.data_end + 1)):
        if rnum in rows and col in rows[rnum]:
            sample.append(rows[rnum][col])

    if not sample:
        return ColRole.UNKNOWN

    str_count = sum(1 for v in sample if isinstance(v, str) and v.strip())
    num_count = sum(1 for v in sample if isinstance(v, (int, float)) and not isinstance(v, bool))
    seq_count = sum(1 for v in sample if _looks_like_sequence(v))

    if str_count > len(sample) * 0.6:
        return ColRole.NAME
    if seq_count > len(sample) * 0.5:
        return ColRole.SEQUENCE
    if num_count > len(sample) * 0.5:
        return ColRole.TOTAL

    return ColRole.UNKNOWN
