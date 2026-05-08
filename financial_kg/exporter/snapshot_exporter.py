"""Snapshot Excel export functionality.

Supports two export modes:
1. values_only: All cells show computed values (no formulas)
2. formula_preserve: Keep formulas, update parameter cells only
"""
from __future__ import annotations
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

from financial_kg.models.graph import FinancialGraph
from financial_kg.engine.snapshot import Snapshot


def parse_cell_id(cell_id: str) -> tuple[str, int, str]:
    """Parse cell_id into (sheet_name, row, col_letter).
    
    Args:
        cell_id: Format like "Sheet1_5_A"
    
    Returns:
        (sheet_name, row_number, column_letter)
    
    Example:
        parse_cell_id("Sheet1_5_A") → ("Sheet1", 5, "A")
    """
    parts = cell_id.split("_")
    if len(parts) < 3:
        raise ValueError(f"Invalid cell_id format: {cell_id}")
    return parts[0], int(parts[1]), parts[2]


def convert_iso_date_to_datetime(value):
    """Convert ISO date string to datetime object for Excel.
    
    Args:
        value: Could be ISO string like "2023-02-01T00:00:00" or other types
    
    Returns:
        datetime object if ISO format detected, otherwise original value
    
    Example:
        "2023-02-01T00:00:00" → datetime(2023, 2, 1)
        "2023-08-31" → "2023-08-31" (no T00:00:00, keep as string)
        123.45 → 123.45 (numeric value unchanged)
    """
    if isinstance(value, str) and "T00:00:00" in value:
        try:
            dt_str = value.replace("T00:00:00", "")
            return datetime.fromisoformat(dt_str)
        except Exception:
            return value
    return value


def validate_template_sheets(template_bytes: bytes, snapshot: Snapshot) -> tuple[bool, list[str]]:
    """Validate if template Excel covers all sheets in snapshot.
    
    Args:
        template_bytes: User-uploaded Excel file bytes
        snapshot: Snapshot object with cell values
    
    Returns:
        (is_valid, missing_sheets): Whether all sheets exist, list of missing sheets
    """
    wb = openpyxl.load_workbook(BytesIO(template_bytes), read_only=True)
    template_sheets = set(wb.sheetnames)
    wb.close()
    
    snapshot_sheets = set()
    for cell_id in snapshot.values.keys():
        sheet_name, _, _ = parse_cell_id(cell_id)
        snapshot_sheets.add(sheet_name)
    
    missing_sheets = [s for s in snapshot_sheets if s not in template_sheets]
    return len(missing_sheets) == 0, missing_sheets


def export_snapshot_to_excel(
    template_bytes: bytes,
    snapshot: Snapshot,
    graph: FinancialGraph,
    mode: str = "values_only",
) -> tuple[bytes, dict]:
    """Export snapshot to Excel with two modes.
    
    Args:
        template_bytes: User-uploaded Excel file bytes (template)
        snapshot: Snapshot object with updated values
        graph: FinancialGraph for cell metadata (formula info)
        mode: Export mode
            - "values_only": All cells show computed values (formula cells replaced)
            - "formula_preserve": Keep formulas, update parameter cells only
    
    Returns:
        (exported_bytes, stats): Excel bytes and export statistics
            stats = {
                "updated_cells": int,
                "formula_preserved": int,
                "skipped_merged": int,
                "missing_sheet_cells": int,
            }
    
    Raises:
        ValueError: If mode is not "values_only" or "formula_preserve"
    """
    if mode not in ("values_only", "formula_preserve"):
        raise ValueError(f"Invalid mode: {mode}. Must be 'values_only' or 'formula_preserve'")
    
    wb = openpyxl.load_workbook(BytesIO(template_bytes), data_only=False)
    
    stats = {
        "updated_cells": 0,
        "formula_preserved": 0,
        "skipped_merged": 0,
        "missing_sheet_cells": 0,
    }
    
    for cell_id, new_value in snapshot.values.items():
        sheet_name, row, col_letter = parse_cell_id(cell_id)
        
        if sheet_name not in wb.sheetnames:
            stats["missing_sheet_cells"] += 1
            continue
        
        ws = wb[sheet_name]
        cell_coord = f"{col_letter}{row}"
        
        graph_cell = graph.cells.get(cell_id)
        
        if graph_cell and graph_cell.is_merged and graph_cell.merge_parent_id:
            stats["skipped_merged"] += 1
            continue
        
        target_cell = ws[cell_coord]
        
        if mode == "values_only":
            converted_value = convert_iso_date_to_datetime(new_value)
            target_cell.value = converted_value
            stats["updated_cells"] += 1
        
        elif mode == "formula_preserve":
            if graph_cell and graph_cell.formula_raw:
                stats["formula_preserved"] += 1
            else:
                converted_value = convert_iso_date_to_datetime(new_value)
                target_cell.value = converted_value
                stats["updated_cells"] += 1
    
    output_buffer = BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    wb.close()
    
    return output_buffer.getvalue(), stats